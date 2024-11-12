# Import package
import openpyxl

# Load workbook
wb = openpyxl.load_workbook("C:\\Users\\lenovo\\OneDrive\\Documents\\TestData.xlsx")
print(wb.sheetnames)
print("Active sheet is " +wb.active.title)

# Create object of any sheet which you want to work on it
sh=wb['Sheet1']
print(sh.title)

# Print the value of selected column
print(sh['A3'].value)

# Create object and get the value
c1 = sh.cell(1,2)
print(c1.value)
print(c1.row)
print(c1.column)

# Find raws having data
rows = sh.max_row
columns = sh.max_column

print(" Total number of rows" + str(rows))
print(" Total number of columns" + str(columns))

for i in range(1,rows+1):
    for j in range(1,columns+1):
        c=sh.cell(i,j)
        print(c.value)